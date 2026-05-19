"""
Extract unique ATS board slugs from a TheirStack 'Known Jobs' Excel export,
then verify every Greenhouse slug against the public Greenhouse Board API.

Outputs:
  - Console summary: ATS breakdown, top slugs by job count
  - greenhouse_verified.csv: greenhouse slugs that returned 200 + jobs
  - greenhouse_dead.csv: greenhouse slugs that 404'd or returned 0 jobs
  - non_greenhouse_slugs.csv: slugs from Ashby/Lever/Workday/etc. (queued for other ingestion paths)
  - unknown_hosts.csv: URLs we couldn't parse (counts, samples)
"""

from __future__ import annotations

import csv
import json
import re
import socket
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path
from typing import Iterator, Optional
from urllib.parse import urlparse, parse_qs

import openpyxl

# INPUT can be a single file (.xlsx or .csv) OR a directory containing
# multiple chunks. Override at the CLI: `python3 extract_slugs.py <path>`.
INPUT_PATH = "/Users/glennfiocca/Desktop/Known Jobs.xlsx"
OUT_DIR = Path("/Users/glennfiocca/Desktop")
GH_API = "https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=false"
ASHBY_API = "https://api.ashbyhq.com/posting-api/job-board/{slug}"
REQUEST_TIMEOUT = 30
REQUEST_DELAY_SEC = 0.15
MAX_RETRIES = 2


ATS_PATTERNS = [
    # (ats_name, hostname_match, slug_regex_against_path)
    ("greenhouse", re.compile(r"(?:^|\.)greenhouse\.io$", re.I), re.compile(r"^/(?:embed/job_app\?for=)?([a-z0-9_-]+)", re.I)),
    ("ashby",      re.compile(r"(?:^|\.)ashbyhq\.com$",  re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("lever",      re.compile(r"(?:^|\.)lever\.co$",     re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("workday",    re.compile(r"myworkday(?:jobs|site)\.com$", re.I), re.compile(r"^/(?:[a-z-]+/)?(?:recruiting/)?([a-z0-9_-]+)", re.I)),
    ("smartrecruiters", re.compile(r"(?:^|\.)smartrecruiters\.com$", re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("icims",      re.compile(r"(?:^|\.)icims\.com$",    re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("recruitee",  re.compile(r"(?:^|\.)recruitee\.com$",re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("teamtailor", re.compile(r"(?:^|\.)teamtailor\.com$", re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
    ("bamboohr",   re.compile(r"(?:^|\.)bamboohr\.com$", re.I), re.compile(r"^/([a-z0-9_-]+)", re.I)),
]


def resolve_grnh(short_url: str) -> Optional[str]:
    """Follow grnh.se redirect to canonical Greenhouse URL. Returns final URL or None."""
    try:
        req = urllib.request.Request(short_url, method="HEAD",
                                     headers={"User-Agent": "launchpad-jobs-token-verifier/1.0"})
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
            return resp.geturl()
    except urllib.error.HTTPError as e:
        # Some shorteners return HTTPError but include the Location header
        loc = e.headers.get("Location") if e.headers else None
        return loc
    except Exception:
        return None


def classify_url(url: str) -> tuple[Optional[str], Optional[str]]:
    """Return (ats_name, slug) or (None, None) if not recognized."""
    try:
        p = urlparse(url)
    except Exception:
        return None, None
    host = (p.hostname or "").lower()
    for ats, host_re, slug_re in ATS_PATTERNS:
        if host_re.search(host):
            # For workday & smartrecruiters, the slug is often a subdomain instead of the path.
            if ats == "workday":
                # host like jobs.<tenant>.wd1.myworkdayjobs.com
                m_host = re.match(r"^(?:[^.]+\.)?([a-z0-9_-]+)\.wd\d+\.myworkdayjobs\.com$", host)
                if m_host:
                    return ats, m_host.group(1)
            if ats == "smartrecruiters":
                m_host = re.match(r"^([a-z0-9_-]+)\.smartrecruiters\.com$", host)
                if m_host and m_host.group(1) not in ("careers", "jobs", "www"):
                    return ats, m_host.group(1)
            if ats == "icims":
                m_host = re.match(r"^([a-z0-9_-]+)\.icims\.com$", host)
                if m_host and m_host.group(1) not in ("careers", "jobs", "www"):
                    return ats, m_host.group(1)
            m = slug_re.match(p.path or "")
            if m:
                return ats, m.group(1).lower()
            return ats, None
    return None, None


def _probe(url: str, jobs_key: str) -> tuple[int, Optional[int]]:
    """GET url, return (http_status, count_or_None). 0 = network failure after retries."""
    req = urllib.request.Request(
        url, headers={"User-Agent": "launchpad-jobs-token-verifier/1.0"},
    )
    for attempt in range(MAX_RETRIES + 1):
        try:
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                data = json.load(resp)
                return resp.status, len(data.get(jobs_key) or [])
        except urllib.error.HTTPError as e:
            return e.code, None
        except (urllib.error.URLError, socket.timeout, TimeoutError,
                json.JSONDecodeError, ConnectionError):
            if attempt < MAX_RETRIES:
                time.sleep(1.0 * (attempt + 1))
                continue
            return 0, None
    return 0, None


def probe_greenhouse(slug: str) -> tuple[int, Optional[int]]:
    return _probe(GH_API.format(slug=slug), "jobs")


def probe_ashby(slug: str) -> tuple[int, Optional[int]]:
    # Ashby's public posting-api returns { "jobs": [...] }.
    return _probe(ASHBY_API.format(slug=slug), "jobs")


REQUIRED_COLUMNS = ["url", "company_name", "job_country_code"]


def discover_input_files(path: str) -> list[Path]:
    """Resolve INPUT_PATH to a list of files. Accepts a single file or a directory."""
    p = Path(path)
    if p.is_file():
        return [p]
    if p.is_dir():
        files = sorted(
            [f for f in p.iterdir()
             if f.is_file() and f.suffix.lower() in (".xlsx", ".csv")]
        )
        if not files:
            raise SystemExit(f"No .xlsx/.csv files found in directory: {path}")
        return files
    raise SystemExit(f"Input path not found: {path}")


def iter_rows_from_file(path: Path) -> Iterator[tuple[dict, list]]:
    """
    Yield (headers_idx_map, row_values) tuples for each data row in the file.
    Supports .xlsx and .csv. Caller does not need to know which.
    """
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        idx = {h: i for i, h in enumerate(headers) if h is not None}
        for r in rows_iter:
            if r is None or all(v is None for v in r):
                continue
            yield idx, list(r)
        wb.close()
        return
    if path.suffix.lower() == ".csv":
        # newline="" so csv handles embedded newlines in quoted fields correctly.
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            idx = {h: i for i, h in enumerate(headers)}
            for r in reader:
                if not r or all(v == "" for v in r):
                    continue
                yield idx, r
        return
    raise SystemExit(f"Unsupported file extension: {path}")


def main() -> int:
    input_path = sys.argv[1] if len(sys.argv) > 1 else INPUT_PATH
    files = discover_input_files(input_path)
    print(f"Input: {input_path}")
    print(f"Found {len(files)} file(s) to process:")
    for f in files:
        print(f"  - {f.name}")
    print()

    # Per ATS: slug -> {companies, countries, sample_url, job_count}
    slugs: dict[tuple[str, str], dict] = {}
    unknown_hosts: dict[str, int] = defaultdict(int)
    unknown_sample: dict[str, str] = {}
    ats_counter: dict[str, int] = defaultdict(int)
    total_rows = 0
    rows_no_url = 0

    # Track companies that show ATS-tracker params (gh_jid/gh_src/ashby_jid)
    # but no direct slug URL. Keyed by (ats, company_name).
    param_only_companies: dict[tuple[str, str], dict] = {}
    # Defer grnh.se resolution until after the first pass.
    grnh_urls: list[tuple[str, str, str]] = []  # (short_url, company, country)

    for file_path in files:
        file_rows = 0
        for idx, r in iter_rows_from_file(file_path):
            total_rows += 1
            file_rows += 1
            # Required columns must be present in the header; missing → skip with hint.
            missing = [c for c in REQUIRED_COLUMNS if c not in idx]
            if missing:
                raise SystemExit(
                    f"{file_path.name}: missing required columns: {missing}. "
                    f"Available: {sorted(idx.keys())}"
                )
            url = r[idx["url"]] if idx["url"] < len(r) else None
            company = r[idx["company_name"]] if idx["company_name"] < len(r) else None
            country = r[idx["job_country_code"]] if idx["job_country_code"] < len(r) else None
            if not url:
                rows_no_url += 1
                continue
            try:
                p = urlparse(url)
                host = (p.hostname or "").lower()
                qs = parse_qs(p.query or "")
            except Exception:
                ats_counter["_unknown"] += 1
                continue

            if host == "grnh.se":
                grnh_urls.append((url, company, country))
                ats_counter["greenhouse_shortlink"] += 1
                continue

            ats, slug = classify_url(url)
            if not ats:
                # Soft signals: ATS tracker params in query string.
                #   gh_jid / gh_src               -> Greenhouse-backed self-hoster
                #   ashby_jid / ashby_embed_*     -> Ashby-backed self-hoster
                param_ats: Optional[str] = None
                if "gh_jid" in qs or "gh_src" in qs:
                    param_ats = "greenhouse"
                elif "ashby_jid" in qs or any(k.startswith("ashby_embed") for k in qs):
                    param_ats = "ashby"
                if param_ats:
                    key = (param_ats, company or "?")
                    entry = param_only_companies.setdefault(key, {
                        "ats": param_ats, "company": company, "rows": 0,
                        "countries": set(), "sample_url": url,
                    })
                    entry["rows"] += 1
                    if country:
                        entry["countries"].add(country)
                    ats_counter[f"{param_ats}_param_only"] += 1
                    continue
                unknown_hosts[host] += 1
                unknown_sample.setdefault(host, url)
                ats_counter["_unknown"] += 1
                continue
            ats_counter[ats] += 1
            if not slug:
                continue
            key = (ats, slug)
            entry = slugs.get(key)
            if entry is None:
                slugs[key] = {
                    "ats": ats,
                    "slug": slug,
                    "companies": {company} if company else set(),
                    "countries": {country} if country else set(),
                    "sample_url": url,
                    "job_rows": 1,
                }
            else:
                if company:
                    entry["companies"].add(company)
                if country:
                    entry["countries"].add(country)
                entry["job_rows"] += 1
        print(f"  - {file_path.name}: {file_rows} rows")

    print()
    print(f"Processed {total_rows} job rows across {len(files)} file(s) ({rows_no_url} had no URL)")
    print()
    print("=== ATS breakdown (rows) ===")
    for ats, n in sorted(ats_counter.items(), key=lambda x: -x[1]):
        print(f"  {ats:<25} {n}")
    print()

    # --- Resolve grnh.se short links to extract canonical Greenhouse slugs ---
    if grnh_urls:
        print(f"=== Resolving {len(grnh_urls)} grnh.se short links ===")
        # Dedupe short URLs first; many job rows share the same short URL? Unlikely, but be safe.
        seen_short: set[str] = set()
        unique_grnh = [(u, c, co) for (u, c, co) in grnh_urls if not (u in seen_short or seen_short.add(u))]
        # To keep runtime sane, sample one per (company, country) — that's all we need for slug discovery.
        by_key: dict[tuple, tuple[str, str, str]] = {}
        for u, c, co in unique_grnh:
            by_key.setdefault((c, co), (u, c, co))
        sampled = list(by_key.values())
        print(f"  ({len(sampled)} unique (company, country) pairs to resolve)")
        for i, (short_url, company, country) in enumerate(sampled, 1):
            resolved = resolve_grnh(short_url)
            if not resolved:
                continue
            ats, slug = classify_url(resolved)
            if ats == "greenhouse" and slug:
                key = ("greenhouse", slug)
                entry = slugs.get(key)
                if entry is None:
                    slugs[key] = {
                        "ats": "greenhouse",
                        "slug": slug,
                        "companies": {company} if company else set(),
                        "countries": {country} if country else set(),
                        "sample_url": resolved,
                        "job_rows": 1,
                    }
                else:
                    if company:
                        entry["companies"].add(company)
                    if country:
                        entry["countries"].add(country)
                    entry["job_rows"] += 1
            if i % 25 == 0:
                print(f"  ...{i}/{len(sampled)}")
            time.sleep(0.1)
        print(f"  done")
        print()

    # Unique slug counts per ATS
    print("=== Unique slugs per ATS ===")
    per_ats: dict[str, list[dict]] = defaultdict(list)
    for entry in slugs.values():
        per_ats[entry["ats"]].append(entry)
    for ats in sorted(per_ats.keys()):
        print(f"  {ats:<20} {len(per_ats[ats])} unique slugs")
    print()

    # Verify slugs per ATS against their public board APIs.
    def verify_set(ats: str, probe_fn, api_label: str) -> tuple[list[dict], list[dict]]:
        ats_slugs = sorted(per_ats.get(ats, []), key=lambda e: -e["job_rows"])
        if not ats_slugs:
            return [], []
        print(f"=== Verifying {len(ats_slugs)} {ats} slugs against {api_label} ===")
        ok_list: list[dict] = []
        dead_list: list[dict] = []
        for i, entry in enumerate(ats_slugs, 1):
            status, jobs = probe_fn(entry["slug"])
            entry["http_status"] = status
            entry["api_jobs"] = jobs
            ok = status == 200 and (jobs or 0) > 0
            marker = "OK " if ok else "DEAD"
            print(f"  [{i:>3}/{len(ats_slugs)}] {marker} {entry['slug']:<35} http={status} jobs={jobs!s:<5} ts_rows={entry['job_rows']}")
            (ok_list if ok else dead_list).append(entry)
            time.sleep(REQUEST_DELAY_SEC)
        return ok_list, dead_list

    verified, dead = verify_set("greenhouse", probe_greenhouse, "boards-api.greenhouse.io")
    ashby_verified, ashby_dead = verify_set("ashby", probe_ashby, "api.ashbyhq.com/posting-api")

    print()
    print(f"Greenhouse verified (live): {len(verified)}")
    print(f"Greenhouse dead/empty:      {len(dead)}")
    print(f"Ashby      verified (live): {len(ashby_verified)}")
    print(f"Ashby      dead/empty:      {len(ashby_dead)}")

    # Write CSVs
    def write(path: Path, rows: list[dict], cols: list[str]) -> None:
        with path.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
            for r in rows:
                w.writerow([
                    "; ".join(sorted(str(x) for x in r[c])) if isinstance(r.get(c), set) else r.get(c, "")
                    for c in cols
                ])

    write(OUT_DIR / "greenhouse_verified.csv", verified,
          ["slug", "companies", "countries", "job_rows", "api_jobs", "sample_url"])
    write(OUT_DIR / "greenhouse_dead.csv", dead,
          ["slug", "companies", "countries", "job_rows", "http_status", "api_jobs", "sample_url"])
    write(OUT_DIR / "ashby_verified.csv", ashby_verified,
          ["slug", "companies", "countries", "job_rows", "api_jobs", "sample_url"])
    write(OUT_DIR / "ashby_dead.csv", ashby_dead,
          ["slug", "companies", "countries", "job_rows", "http_status", "api_jobs", "sample_url"])

    other_ats = [e for e in slugs.values() if e["ats"] not in ("greenhouse", "ashby")]
    write(OUT_DIR / "other_ats_slugs.csv", other_ats,
          ["ats", "slug", "companies", "countries", "job_rows", "sample_url"])

    with (OUT_DIR / "unknown_hosts.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["host", "row_count", "sample_url"])
        for host, n in sorted(unknown_hosts.items(), key=lambda x: -x[1]):
            w.writerow([host, n, unknown_sample.get(host, "")])

    # Companies with ATS tracker params but no direct slug URL.
    # Split by ATS so the user can hand each list to the right derive-and-verify pass.
    with (OUT_DIR / "param_only_companies.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["ats", "company", "rows", "countries", "sample_url"])
        for entry in sorted(param_only_companies.values(), key=lambda e: -e["rows"]):
            w.writerow([
                entry["ats"], entry["company"], entry["rows"],
                "; ".join(sorted(entry["countries"])),
                entry["sample_url"],
            ])
    print(f"Wrote: {OUT_DIR / 'param_only_companies.csv'} ({len(param_only_companies)} entries)")

    print()
    print(f"Wrote: {OUT_DIR / 'greenhouse_verified.csv'}")
    print(f"Wrote: {OUT_DIR / 'greenhouse_dead.csv'}")
    print(f"Wrote: {OUT_DIR / 'ashby_verified.csv'}")
    print(f"Wrote: {OUT_DIR / 'ashby_dead.csv'}")
    print(f"Wrote: {OUT_DIR / 'other_ats_slugs.csv'}")
    print(f"Wrote: {OUT_DIR / 'unknown_hosts.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
